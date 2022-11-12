import os.path
import csv
from PyPDF2 import PdfReader
from reportlab.pdfgen.canvas import Canvas
import zipfile
from zipfile import ZipFile
import openpyxl
from openpyxl import load_workbook
import shutil
import pathlib
from pathlib import Path

path = pathlib.Path.cwd()
print(path)

#создаем PDF файл
canvas = Canvas("pdf_unit_6.pdf")
canvas.drawString(1, 1, "eto_PDF")
canvas.save()

#создаем XLSX файл
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row=1, column=1)
c1.value = "ДАТА"
wb.save(filename='xlsx_unit_6.xlsx')

#создаем CSV файл
with open('Unit_6.csv', "w", newline='') as csv_file:
    writer = csv.writer(csv_file, delimiter=' ')
    writer.writerow(["РЕЙТИНГ"])

# создаем архив
shutil.make_archive('archive', 'zip', os.chdir("./resourses"))

#добавляем файлы в архив
zipFile = zipfile.ZipFile('archive.zip', 'w', zipfile.ZIP_DEFLATED)
path_pdf = Path(path,'pdf_unit_6.pdf')
path_xlsx = Path(path,'xlsx_unit_6.xlsx')
path_csv = Path(path,'Unit_6.csv')
zipFile.write(path_pdf, 'my.pdf')
zipFile.write(path_xlsx, 'my.xlsx')
zipFile.write(path_csv, 'my_csv.csv')
zipFile.close()

#извлекаем
with ZipFile('archive.zip') as myzip:
    myzip.extractall()


# #открываем и проверяем PDF
#
pdf_reader = PdfReader('my.pdf')
number_of_pages = len(pdf_reader.pages)
assert number_of_pages == 1
page = pdf_reader.pages[0]
text = page.extract_text()
assert 'eto_PDF' in text

#открываем и проверяем XLSX файл
workbook = load_workbook('my.xlsx')
sheet = workbook.active
assert 'ДАТА' in sheet.cell(row=1, column=1).value

#открываем и проверяем CSV файл
with open('my_csv.csv') as csvfile:
    table = csv.reader(csvfile)
    for row in table:
        assert 'РЕЙТИНГ' in row